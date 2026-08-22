"""Parsing and AI-relevance scoring for the Aizentify applications export.

Reads the Tally form export (CSV or XLSX, same template each time), infers
each applicant's Indian state from their free-text city, scores AI/Claude
relevance from their certifications and employer text, classifies them into
an Intern or Experienced track, and returns one typed row per candidate who
is willing to relocate to Chennai. Runs once per upload -- the caller
persists the result and never re-runs this on view.
"""

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

COL_SUBMISSION_ID = "Submission ID"
COL_SUBMITTED_AT = "Submitted at"
COL_ROLE = "Role Applying For"
COL_NAME = "Name"
COL_EMAIL = "Email"
COL_MOBILE = "Mobile Number"
COL_EMPLOYER = "Current Employer"
COL_EXPERIENCE = "Total Professional Experience"
COL_LOCATION = "Your Current Location"
COL_CERTS = "Relevant Certifications if any (Claude Certified Preferred)"
COL_RELOCATE = "Willing to relocate to Chennai?"
COL_RESUME = "Upload Resume"

AI_KEYWORDS = [
    "claude", "anthropic", "genai", "gen ai", "generative ai", "generative-ai",
    "llm", "large language model", "rag", "retrieval augmented", "machine learning",
    " ml ", "ml)", "ml,", "artificial intelligence", "deep learning", "nlp",
    "natural language", "openai", "gpt", "chatgpt", "agentic", "ai agent",
    "computer vision", "neural network", "prompt engineering", "data science",
    "databricks", "hugging face", "huggingface", "tensorflow", "pytorch",
    "langchain", "vector database", "vector db", "transformer",
]
CLAUDE_KEYWORDS = ["claude", "anthropic"]
EMPLOYER_AI_KEYWORDS = ["ai", "ml", "data", "tech", "software", "analytics", "intelligence"]

STATE_KEYWORDS = [
    ("andhra pradesh", "Andhra Pradesh"),
    ("andhrapradesh", "Andhra Pradesh"),
    ("madhya pradesh", "Madhya Pradesh"),
    ("uttar pradesh", "Uttar Pradesh"),
    ("west bengal", "West Bengal"),
    ("tamil nadu", "Tamil Nadu"),
    ("tamilnadu", "Tamil Nadu"),
    ("karnataka", "Karnataka"),
    ("karnatak", "Karnataka"),
    ("kerala", "Kerala"),
    ("telangana", "Telangana"),
    ("gujarat", "Gujarat"),
    ("rajasthan", "Rajasthan"),
    ("maharashtra", "Maharashtra"),
    ("chandigarh", "Chandigarh"),
    ("puducherry", "Puducherry"),
    ("pondicherry", "Puducherry"),
    ("new delhi", "Delhi"),
    ("delhi", "Delhi"),
    (", up", "Uttar Pradesh"),
    (" up,", "Uttar Pradesh"),
    (" mp,", "Madhya Pradesh"),
    (", mp", "Madhya Pradesh"),
]

CITY_KEYWORDS = [
    ("visakhapatnam", "Andhra Pradesh"),
    ("tadepalligudem", "Andhra Pradesh"),
    ("veeravasram", "Andhra Pradesh"),
    ("west godavari", "Andhra Pradesh"),
    ("bhimavaram", "Andhra Pradesh"),
    ("bhimavarm", "Andhra Pradesh"),
    ("nellore", "Andhra Pradesh"),
    ("tirupati", "Andhra Pradesh"),
    ("guntur", "Andhra Pradesh"),
    ("greater noida", "Uttar Pradesh"),
    ("saravanampatti", "Tamil Nadu"),
    ("thoraipakkam", "Tamil Nadu"),
    ("tirupattur", "Tamil Nadu"),
    ("chengalpattu", "Tamil Nadu"),
    ("krishnagiri", "Tamil Nadu"),
    ("dharmapuri", "Tamil Nadu"),
    ("ulundurpet", "Tamil Nadu"),
    ("dindigul", "Tamil Nadu"),
    ("coimbatore", "Tamil Nadu"),
    ("namakkal", "Tamil Nadu"),
    ("tiruppur", "Tamil Nadu"),
    ("madurai", "Tamil Nadu"),
    ("chennai", "Tamil Nadu"),
    ("hosur", "Tamil Nadu"),
    ("salem", "Tamil Nadu"),
    ("erode", "Tamil Nadu"),
    ("ooty", "Tamil Nadu"),
    ("chikkamagalur", "Karnataka"),
    ("davanagere", "Karnataka"),
    ("davangere", "Karnataka"),
    ("bangalore", "Karnataka"),
    ("banglore", "Karnataka"),
    ("bengaluru", "Karnataka"),
    ("bengalur", "Karnataka"),
    ("malavalli", "Karnataka"),
    ("bagalkot", "Karnataka"),
    ("dharwad", "Karnataka"),
    ("hospet", "Karnataka"),
    ("mandya", "Karnataka"),
    ("mysore", "Karnataka"),
    ("hubli", "Karnataka"),
    ("udupi", "Karnataka"),
    ("bidar", "Karnataka"),
    ("kasaragod", "Kerala"),
    ("ernakulam", "Kerala"),
    ("thrissur", "Kerala"),
    ("palakkad", "Kerala"),
    ("calicut", "Kerala"),
    ("kochi", "Kerala"),
    ("bhadrachalam", "Telangana"),
    ("karimnagar", "Telangana"),
    ("hyderabad", "Telangana"),
    ("hydrabad", "Telangana"),
    ("ahmedabad", "Gujarat"),
    ("rajkot", "Gujarat"),
    ("ghaziabad", "Uttar Pradesh"),
    ("gorakhpur", "Uttar Pradesh"),
    ("aligarh", "Uttar Pradesh"),
    ("lucknow", "Uttar Pradesh"),
    ("noida", "Uttar Pradesh"),
    ("agra", "Uttar Pradesh"),
    ("bhopal", "Madhya Pradesh"),
    ("indore", "Madhya Pradesh"),
    ("sehore", "Madhya Pradesh"),
    ("ashta", "Madhya Pradesh"),
    ("nagpur", "Maharashtra"),
    ("mumbai", "Maharashtra"),
    ("pune", "Maharashtra"),
    ("jaipur", "Rajasthan"),
    ("falna", "Rajasthan"),
    ("kolkata", "West Bengal"),
]


def infer_state(city_raw: str) -> str:
    if not city_raw or city_raw.strip().lower() in ("not specified", "na", "n/a", "-", "remote"):
        return "Not specified"
    c = city_raw.strip().lower()
    for kw, state in STATE_KEYWORDS:
        if kw in c:
            return state
    for kw, state in CITY_KEYWORDS:
        if kw in c:
            return state
    return "Not specified"


def parse_exp_years(raw: str) -> float:
    if not raw:
        return 0.0
    s = raw.strip().lower()
    if not s or s in ("na", "n/a", "fresher", "-"):
        return 0.0
    m = re.search(r"(\d+(\.\d+)?)", s)
    if not m:
        return 0.0
    val = float(m.group(1))
    if "month" in s and "year" not in s:
        val = val / 12.0
    return val


def score_ai_signal(cert_text: str, employer_text: str):
    text = (cert_text or "").lower()
    score = 0
    matched = set()
    for kw in AI_KEYWORDS:
        if kw.strip() in text:
            score += 1
            matched.add(kw.strip())
    claude_bonus = 0
    for kw in CLAUDE_KEYWORDS:
        if kw in text:
            claude_bonus += 3
    emp = (employer_text or "").lower()
    emp_bonus = 1 if any(k in emp for k in EMPLOYER_AI_KEYWORDS) else 0
    return score + claude_bonus + emp_bonus, sorted(matched), claude_bonus > 0


def classify_tier(role: str) -> str:
    if "Intern" in role:
        return "Intern"
    if "Full Stack" in role or "Associate" in role:
        return "Experienced"
    return "Other"


def read_rows(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    if suffix in (".xlsx", ".xlsm"):
        ws = load_workbook(path, read_only=True, data_only=True).active
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in raw_row):
                continue
            rows.append({h: ("" if v is None else str(v)) for h, v in zip(header, raw_row) if h})
        return rows
    raise ValueError(f"Unsupported file type: {suffix} (use .csv or .xlsx)")


def build_report(path: Path) -> dict:
    """Parse an upload and score every relocating candidate. Runs once, at upload time."""
    raw_rows = read_rows(path)
    relocate_yes = [r for r in raw_rows if (r.get(COL_RELOCATE) or "").strip().lower() == "yes"]

    candidates = []
    for i, r in enumerate(relocate_yes):
        role = (r.get(COL_ROLE) or "").strip()
        cert = (r.get(COL_CERTS) or "").strip()
        employer = (r.get(COL_EMPLOYER) or "").strip()
        city = (r.get(COL_LOCATION) or "").strip() or "Not specified"
        exp_years = parse_exp_years(r.get(COL_EXPERIENCE) or "")
        ai_score, matched_keywords, claude_mentioned = score_ai_signal(cert, employer)
        tier = classify_tier(role)

        final_score = ai_score * 2 if tier == "Intern" else ai_score * 2 + min(exp_years, 6) * 1.5

        submission_id = (r.get(COL_SUBMISSION_ID) or "").strip()
        uid = submission_id or f"row-{i}"

        candidates.append({
            "uid": uid,
            "name": (r.get(COL_NAME) or "").strip(),
            "email": (r.get(COL_EMAIL) or "").strip(),
            "mobile": (r.get(COL_MOBILE) or "").strip(),
            "city": city,
            "state": infer_state(city),
            "role": role,
            "tier": tier,
            "exp_raw": (r.get(COL_EXPERIENCE) or "").strip(),
            "exp_years": round(exp_years, 2),
            "employer": employer or "Not specified",
            "cert_text": cert,
            "ai_score": ai_score,
            "claude_mentioned": claude_mentioned,
            "matched_keywords": matched_keywords,
            "final_score": round(final_score, 2),
            "resume": (r.get(COL_RESUME) or "").strip(),
            "submitted_at": (r.get(COL_SUBMITTED_AT) or "").strip(),
            "rating": 0,
            "round1": False,
            "round2": False,
        })

    return {
        "candidates": candidates,
        "total_submissions": len(raw_rows),
        "relocate_yes": len(relocate_yes),
        "relocate_no": len(raw_rows) - len(relocate_yes),
    }
