"""
Bulk-uploads candidates for a round from a CSV or XLSX file and creates one
Attempt per candidate, tracking who's invited so responses can be matched
back to the roster later during scoring.

Required columns: name, email (or "Personal Email ID" - see HEADER_ALIASES).
Optional profile columns (stored if present, ignored if not): reg no,
gender, degree, stream, resume, github url, portfolio url.

The college's own placement-list export can usually be uploaded as-is -
HEADER_ALIASES maps its real column names (e.g. "Reg No.", "Personal Email
ID", "  GitHub Profile URL  ", "Project / Portfolio URL (if available)")
to the fields above, so no manual renaming is needed. "Personal Email ID"
is deliberately mapped to "email" (not "College Email ID"): invites and
Google Forms need the candidate's real Gmail for Google auth, which a
college email address usually isn't tied to.

Usage:
    python add_candidates.py --round-id 1 --file candidates.csv
    python add_candidates.py --round-id 1 --file candidates.xlsx
"""

import argparse
import csv
import re
import secrets
from pathlib import Path

from openpyxl import load_workbook

from models import Attempt, Candidate, Round, SessionLocal, init_db

PROFILE_FIELDS = ("reg_no", "gender", "degree", "stream", "resume_url", "github_url", "portfolio_url")

HEADER_ALIASES = {
    "name": "name",
    "email": "email",
    "personal email id": "email",
    "personal email": "email",
    "reg no": "reg_no",
    "regno": "reg_no",
    "registration no": "reg_no",
    "gender": "gender",
    "degree": "degree",
    "stream": "stream",
    "resume": "resume_url",
    "resume url": "resume_url",
    "resume link": "resume_url",
    "github profile url": "github_url",
    "github url": "github_url",
    "github": "github_url",
    "project / portfolio url": "portfolio_url",
    "portfolio url": "portfolio_url",
    "portfolio": "portfolio_url",
}


def _normalize_header(h) -> str:
    if not isinstance(h, str):
        return h
    h = h.strip().lower().replace(".", "")
    h = re.sub(r"\(.*?\)", "", h)  # drop parenthetical notes like "(if available)"
    return re.sub(r"\s+", " ", h).strip()


def rows_from_file(path: str) -> list[dict]:
    file_path = Path(path)
    if file_path.suffix.lower() == ".csv":
        rows = list(csv.DictReader(file_path.open(newline="", encoding="utf-8")))
    elif file_path.suffix.lower() == ".xlsx":
        ws = load_workbook(file_path, read_only=True, data_only=True).active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in raw_row):
                continue
            rows.append(dict(zip(header, raw_row)))
    else:
        raise ValueError(f"Unsupported file type: {file_path.suffix} (use .csv or .xlsx)")

    # Column names are matched case-insensitively and via HEADER_ALIASES, so
    # both our own simple template and a real college export (with headers
    # like "Reg No." or "Personal Email ID") normalize to the same internal
    # keys - downstream code can always just use row["email"]/row["name"]/etc.
    rows = [
        {HEADER_ALIASES.get(_normalize_header(k), _normalize_header(k)): v for k, v in row.items()}
        for row in rows
    ]

    found = set(rows[0].keys()) if rows else set()
    missing = {"name", "email"} - found
    if missing:
        hint = ""
        if len(found) == 1 and "," in next(iter(found), ""):
            hint = (
                " It looks like the whole header line is wrapped in one pair of "
                "quotes (e.g. \"name,email\"), so it's being read as a single "
                "column instead of two - remove the surrounding quotes and "
                "re-save as plain comma-separated values."
            )
        raise ValueError(
            f"Missing required column(s) {sorted(missing)} in {path}. "
            f"Found columns: {sorted(found)}.{hint}"
        )
    return rows


def add_candidates_from_file(session, round_: Round, file_path: str) -> tuple[int, int]:
    """Returns (created, already_invited) - already_invited counts rows that
    matched an existing Attempt for this round (no duplicate invite/token
    created, by design) but whose profile fields were still refreshed."""
    created = 0
    already_invited = 0
    for row in rows_from_file(file_path):
        email = str(row["email"]).strip().lower()
        name = str(row["name"]).strip()

        candidate = (
            session.query(Candidate)
            .filter_by(drive_id=round_.drive_id, email=email)
            .first()
        )
        if candidate is None:
            candidate = Candidate(drive_id=round_.drive_id, email=email, name=name)
            session.add(candidate)
            session.flush()

        # Fills in profile data whenever a non-empty value is supplied -
        # covers both first upload and a later, fuller re-upload for the
        # same drive - without ever clobbering a known value with a blank.
        for field in PROFILE_FIELDS:
            value = row.get(field)
            if value is not None and str(value).strip():
                setattr(candidate, field, str(value).strip())

        existing = (
            session.query(Attempt)
            .filter_by(candidate_id=candidate.id, round_id=round_.id)
            .first()
        )
        if existing is not None:
            already_invited += 1
            continue

        session.add(
            Attempt(
                candidate_id=candidate.id,
                round_id=round_.id,
                token=secrets.token_urlsafe(24),
                status="invited",
            )
        )
        created += 1

    session.commit()
    return created, already_invited


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--round-id", type=int, required=True)
    parser.add_argument("--file", required=True, help="candidates .csv or .xlsx")
    args = parser.parse_args()

    init_db()
    session = SessionLocal()
    round_ = session.get(Round, args.round_id)
    if round_ is None:
        raise SystemExit(f"No round with id {args.round_id}")

    try:
        created, already_invited = add_candidates_from_file(session, round_, args.file)
    except ValueError as e:
        raise SystemExit(str(e))
    print(f"Created {created} attempt(s) for round #{round_.id}", end="")
    print(f" ({already_invited} already invited - profile data refreshed only)" if already_invited else "")


if __name__ == "__main__":
    main()
