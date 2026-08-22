"""
Generates question_bank_v1.xlsx: a real, filled 50-question starter bank
(25 Aptitude for 4th-year engineering internship screening, 25
Programming covering Python, TypeScript, LLM basics, and Claude
Platform foundation-readiness). Uses the same 3-sheet schema as
make_template.py (Aptitude / Programming / Answer Key).

Usage:
    python build_initial_bank.py
    python load_questions.py question_bank_v1.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

QUESTION_HEADERS = ["question_id", "topic", "difficulty", "text", "option_a", "option_b", "option_c", "option_d", "points"]
ANSWER_KEY_HEADERS = ["question_id", "correct_option"]

# (question_id, topic, difficulty, text, A, B, C, D, points, correct)
APTITUDE = [
    ("APT001", "quant", "easy", "A shopkeeper marks up an item by 25% and then gives a 20% discount on the marked price. What is the net profit/loss?", "0%", "5% loss", "5% profit", "10% loss", 1, "A"),
    ("APT002", "quant", "easy", "The average of 5 numbers is 20. If one number is excluded, the average becomes 18. What is the excluded number?", "24", "26", "28", "30", 1, "C"),
    ("APT003", "quant", "medium", "A can complete a work in 12 days, B in 18 days. Working together, how many days will they take?", "7 days", "7.2 days", "8 days", "9 days", 1, "B"),
    ("APT004", "quant", "easy", "The ratio of two numbers is 3:5 and their sum is 96. Find the larger number.", "36", "54", "60", "66", 1, "C"),
    ("APT005", "quant", "easy", "A train 150 m long crosses a pole in 15 seconds. What is its speed in km/h?", "30", "33", "36", "40", 1, "C"),
    ("APT006", "quant", "medium", "The simple interest on a sum for 2 years at 10% p.a. is Rs. 400. Find the principal.", "Rs. 1800", "Rs. 2000", "Rs. 2200", "Rs. 2500", 1, "B"),
    ("APT007", "quant", "easy", "What is 40% of 250 minus 15% of 200?", "60", "65", "70", "75", 1, "C"),
    ("APT008", "quant", "medium", "Find the compound interest on Rs. 10,000 for 2 years at 10% p.a. compounded annually.", "Rs. 2000", "Rs. 2100", "Rs. 2200", "Rs. 2500", 1, "B"),
    ("APT009", "logical", "medium", "Find the next number in the series: 2, 6, 12, 20, 30, ?", "40", "42", "44", "36", 1, "B"),
    ("APT010", "logical", "medium", "If CODING is written as DPEJOH (each letter shifted forward by 1), how is FLOWER written?", "GMPXFS", "GMPXES", "FMPXFS", "GLPXFS", 1, "A"),
    ("APT011", "logical", "easy", "Find the odd one out: Triangle, Square, Circle, Cube", "Triangle", "Square", "Circle", "Cube", 1, "D"),
    ("APT012", "logical", "medium", "All pens are pencils. All pencils are erasers. Which conclusion follows?", "All pens are erasers", "All erasers are pens", "No pens are erasers", "Cannot be determined", 1, "A"),
    ("APT013", "logical", "easy", "Find the odd one out: 8, 27, 64, 100, 125", "8", "100", "64", "125", 1, "B"),
    ("APT014", "logical", "medium", "Pointing to a photograph, a man says, 'She is the daughter of my grandfather's only son.' How is the girl related to the man?", "Daughter", "Sister", "Mother", "Niece", 1, "B"),
    ("APT015", "logical", "medium", "In a code language, TRAIN is written as USBJO (each letter shifted forward by 1). How is BUS written?", "CVT", "CVU", "BVT", "CVS", 1, "A"),
    ("APT016", "logical", "easy", "Complete the analogy: Doctor : Hospital :: Teacher : ?", "Book", "School", "Student", "Class", 1, "B"),
    ("APT017", "verbal", "easy", "Choose the word closest in meaning to 'Meticulous'.", "Careless", "Precise", "Lazy", "Quick", 1, "B"),
    ("APT018", "verbal", "easy", "Choose the word most opposite in meaning to 'Benevolent'.", "Kind", "Generous", "Cruel", "Charitable", 1, "C"),
    ("APT019", "verbal", "easy", "Choose the best word to complete: 'Despite the heavy rain, the match was ___ played.'", "hardly", "barely", "still", "rarely", 1, "C"),
    ("APT020", "verbal", "easy", "Choose the correctly spelled word.", "Recieve", "Receive", "Receeve", "Receve", 1, "B"),
    ("APT021", "verbal", "medium", "Choose the best word to complete: 'The manager was ___ by the team's poor performance.'", "elated", "disappointed", "amused", "indifferent", 1, "B"),
    ("APT022", "data_interpretation", "medium", "In a class of 80 students, 45 play cricket, 30 play football, and 10 play both. How many play neither?", "10", "15", "20", "25", 1, "B"),
    ("APT023", "data_interpretation", "easy", "A company's revenue grew from Rs. 50 lakh to Rs. 65 lakh in one year. What is the percentage growth?", "20%", "25%", "30%", "35%", 1, "C"),
    ("APT024", "data_interpretation", "easy", "Out of 200 candidates who appeared for a test, 60% passed. How many failed?", "60", "70", "80", "90", 1, "C"),
    ("APT025", "data_interpretation", "medium", "The ratio of boys to girls in a college is 5:4 and there are 180 girls. How many boys are there?", "200", "210", "225", "240", 1, "C"),
]

PROGRAMMING = [
    ("PROG001", "python", "easy", "Which keyword is used to define a function in Python?", "function", "def", "func", "define", 1, "B"),
    ("PROG002", "python", "easy", "What is the output of: print(type([]))", "<class 'list'>", "<class 'tuple'>", "<class 'dict'>", "<class 'set'>", 1, "A"),
    ("PROG003", "python", "easy", "Which method adds a single item to the end of a Python list?", "append()", "add()", "insert()", "extend()", 1, "A"),
    ("PROG004", "python", "easy", "What does len({'a': 1, 'b': 2, 'c': 3}) return?", "1", "2", "3", "Error", 1, "C"),
    ("PROG005", "python", "easy", "Which keyword is used to catch exceptions in Python?", "catch", "except", "rescue", "handle", 1, "B"),
    ("PROG006", "typescript", "easy", "Which keyword defines a block-scoped constant in TypeScript?", "var", "let", "const", "static", 1, "C"),
    ("PROG007", "typescript", "medium", "Which TypeScript type disables type checking and allows a value to be of any type?", "unknown", "any", "never", "void", 1, "B"),
    ("PROG008", "typescript", "easy", "What is the correct syntax to define an interface in TypeScript?", "interface Person { name: string }", "class Person { name: string }", "struct Person { name: string }", "type Person => { name: string }", 1, "A"),
    ("PROG009", "typescript", "easy", "Which symbol marks a property as optional in a TypeScript interface?", "!", "?", "*", "&", 1, "B"),
    ("PROG010", "typescript", "easy", "What does running the 'tsc' command do?", "Runs TypeScript code directly", "Compiles TypeScript to JavaScript", "Installs TypeScript packages", "Lints TypeScript code", 1, "B"),
    ("PROG011", "llm_basics", "easy", "What does 'LLM' stand for?", "Large Language Model", "Linear Learning Machine", "Long Loop Memory", "Logic Language Module", 1, "A"),
    ("PROG012", "llm_basics", "easy", "In the context of LLMs, what is 'tokenization'?", "Encrypting the input text", "Breaking text into smaller units for processing", "Compressing the model's size", "Training the model on new data", 1, "B"),
    ("PROG013", "llm_basics", "easy", "What is a 'prompt' in the context of an LLM?", "The model's training dataset", "The input text given to the model to generate a response", "The output generated by the model", "The model's internal weights", 1, "B"),
    ("PROG014", "llm_basics", "medium", "What does 'context window' refer to in an LLM?", "The physical hardware running the model", "The maximum amount of text (in tokens) the model can consider at once", "The programming interface to the model", "The time taken to generate a response", 1, "B"),
    ("PROG015", "llm_basics", "medium", "What is 'fine-tuning' in the context of LLMs?", "Adjusting the model's temperature setting", "Further training a pre-trained model on a specific dataset to specialize it", "Reducing the model's context window", "Compressing the model for faster inference", 1, "B"),
    # The 10 questions below are mapped to the "Claude Certified Associate"
    # exam blueprint domains, weighted proportionally to the domain's
    # exam weighting (21/16/15/14/12/12/10 % -> 2/2/2/1/1/1/1 questions).
    ("PROG016", "output_evaluation_validation", "medium", "An AI assistant gives a confident-sounding answer that includes a fact which turns out not to be true or verifiable. This is known as a:", "Token limit", "Hallucination", "Context window", "Temperature error", 1, "B"),
    ("PROG017", "output_evaluation_validation", "medium", "Before relying on an AI-generated answer for an important business decision, what is the best practice?", "Accept it immediately since AI is always accurate", "Verify key facts/citations and escalate for human review if uncertain", "Ignore AI outputs entirely", "Only check the first sentence", 1, "B"),
    ("PROG018", "workflow_integration_design", "medium", "Which of the following best illustrates integrating Claude into an existing business workflow?", "Using Claude once, manually, with no connection to other tools", "Embedding Claude into a recurring task (e.g., drafting weekly reports) with consistent instructions and inputs", "Asking Claude general trivia questions", "Avoiding Claude for any repeated tasks", 1, "B"),
    ("PROG019", "workflow_integration_design", "medium", "When designing a solution that uses Claude as part of a marketing content pipeline, what should be defined first?", "The color scheme of the website", "The specific task, inputs, and desired output format for each step", "Only the AI's response length", "Nothing, Claude works the same for every task", 1, "B"),
    ("PROG020", "governance_risk_responsible_use", "easy", "Before sharing sensitive customer or company data with an AI assistant, an employee should:", "Share freely since AI conversations are always private", "Check the organization's acceptable use and data privacy policies first", "Only share it with the AI, never with coworkers", "Assume there are no risks", 1, "B"),
    ("PROG021", "governance_risk_responsible_use", "medium", "What is a key part of 'responsible use' when deploying an AI assistant in a business setting?", "Using it for any task without oversight", "Adhering to acceptable use policies and avoiding exposure of sensitive information", "Sharing all outputs publicly", "Disabling all human review", 1, "B"),
    ("PROG022", "prompting_task_execution", "medium", "A well-structured prompt commonly includes which of the following elements?", "Role, context, task, constraints, and desired format", "Only a single keyword", "Random unrelated examples", "No instructions at all, just a question", 1, "A"),
    ("PROG023", "product_model_selection", "easy", "For a simple, high-volume, low-complexity task, which type of model tier is typically best suited (for cost and speed)?", "The largest, most expensive model available", "A faster, lighter-weight model tier", "It doesn't matter which model is used", "A model with no context window", 1, "B"),
    ("PROG024", "config_knowledge_management", "medium", "In a project-based workspace feature (like Claude Projects), what is the purpose of adding custom instructions and reference documents?", "To slow down responses intentionally", "To give the assistant consistent context and guidance so it produces relevant, on-brand outputs for that project", "To permanently change the model's core training", "To restrict the assistant from answering any questions", 1, "B"),
    ("PROG025", "troubleshooting_optimization", "easy", "If an AI assistant's response to a task is consistently too vague or off-target, what is the best first step to improve it?", "Give up and do the task manually every time", "Refine the prompt with more specific context, constraints, and examples", "Repeat the exact same prompt again", "Reduce the amount of information given", 1, "B"),
]


def add_validation(ws, headers, max_row=500):
    if "difficulty" in headers:
        col = get_column_letter(headers.index("difficulty") + 1)
        dv = DataValidation(type="list", formula1='"easy,medium,hard"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max_row}")
    if "correct_option" in headers:
        col = get_column_letter(headers.index("correct_option") + 1)
        dv = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max_row}")


def build_question_sheet(wb, name, rows):
    ws = wb.create_sheet(name)
    ws.append(QUESTION_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(list(row[:9]))
    add_validation(ws, QUESTION_HEADERS)
    for i, width in enumerate([12, 18, 10, 65, 20, 20, 20, 20, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_answer_key_sheet(wb, all_rows):
    ws = wb.create_sheet("Answer Key")
    ws.append(ANSWER_KEY_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in all_rows:
        ws.append([row[0], row[9]])
    add_validation(ws, ANSWER_KEY_HEADERS)
    for i, width in enumerate([12, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    build_question_sheet(wb, "Aptitude", APTITUDE)
    build_question_sheet(wb, "Programming", PROGRAMMING)
    build_answer_key_sheet(wb, APTITUDE + PROGRAMMING)

    out_path = Path(__file__).parent / "question_bank_v1.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}: {len(APTITUDE)} aptitude + {len(PROGRAMMING)} programming questions")


if __name__ == "__main__":
    main()
