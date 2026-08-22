"""
Generates question_bank_template.xlsx: an editable workbook with one tab
per round type (Aptitude, Programming) plus a separate "Answer Key" tab.
Question sheets never carry the correct answer, only a question_id -
keeps the reviewable question sheet clean of answers. Fill it in and hand
it to load_questions.py whenever the bank needs updating.

Usage:
    python make_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

QUESTION_HEADERS = [
    "question_id",
    "topic",
    "difficulty",
    "text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "points",
]
QUESTION_WIDTHS = [12, 12, 10, 60, 16, 16, 16, 16, 8]

ANSWER_KEY_HEADERS = ["question_id", "correct_option"]
ANSWER_KEY_WIDTHS = [12, 14]

EXAMPLES = {
    "Aptitude": (
        ["APT001", "quant", "easy", "A train travels 60 km in 45 minutes. What is its speed in km/h?", "60", "80", "75", "90", 1],
        ("APT001", "C"),
    ),
    "Programming": (
        ["PROG001", "dsa", "easy", "What is the time complexity of binary search on a sorted array?", "O(n)", "O(log n)", "O(n^2)", "O(1)", 1],
        ("PROG001", "B"),
    ),
}


def add_validation(ws, headers, max_row=500):
    if "difficulty" in headers:
        col = get_column_letter(headers.index("difficulty") + 1)
        dv = DataValidation(type="list", formula1='"easy,medium,hard"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max_row}")
    if "correct_option" in headers:
        col = get_column_letter(headers.index("correct_option") + 1)
        dv = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max_row}")


def build_question_sheet(wb, name, example_row):
    ws = wb.create_sheet(name)
    ws.append(QUESTION_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.append(example_row)
    add_validation(ws, QUESTION_HEADERS)
    for i, width in enumerate(QUESTION_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_answer_key_sheet(wb, example_pairs):
    ws = wb.create_sheet("Answer Key")
    ws.append(ANSWER_KEY_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for qid, correct in example_pairs:
        ws.append([qid, correct])
    add_validation(ws, ANSWER_KEY_HEADERS)
    for i, width in enumerate(ANSWER_KEY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    example_pairs = []
    for name, (example_row, answer_pair) in EXAMPLES.items():
        build_question_sheet(wb, name, example_row)
        example_pairs.append(answer_pair)
    build_answer_key_sheet(wb, example_pairs)

    out_path = Path(__file__).parent / "question_bank_template.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
