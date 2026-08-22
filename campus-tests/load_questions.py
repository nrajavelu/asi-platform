"""
Loads the question bank into the local SQLite DB. Accepts either:
  - a .csv file with "round_type" and "correct_option" columns, or
  - a .xlsx workbook with one sheet per round type ("Aptitude",
    "Programming") holding a "question_id" column but no answer, plus a
    separate "Answer Key" sheet mapping question_id -> correct_option
    (see make_template.py). Keeping answers off the question sheets lets
    a question sheet be shared/reviewed without revealing correct answers.

The file is treated as the source of truth for the bank: each run
replaces all rows.

Usage:
    python load_questions.py question_bank_template.csv
    python load_questions.py question_bank_template.xlsx
"""

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

from models import Question, SessionLocal, init_db

CSV_COLUMNS = [
    "topic",
    "difficulty",
    "text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_option",
    "points",
]

XLSX_QUESTION_COLUMNS = [
    "question_id",
    "topic",
    "difficulty",
    "text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "points",
]

SHEET_TO_ROUND_TYPE = {"Aptitude": "aptitude", "Programming": "programming"}
ANSWER_KEY_SHEET = "Answer Key"


def rows_from_csv(path: Path) -> list[dict]:
    rows = []
    for row in csv.DictReader(path.open(newline="", encoding="utf-8")):
        rows.append({key: row[key] for key in CSV_COLUMNS} | {"round_type": row["round_type"]})
    return rows


def _read_sheet(ws) -> list[dict]:
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in raw_row):
            continue
        records.append(dict(zip(header, raw_row)))
    return records


def rows_from_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if ANSWER_KEY_SHEET not in wb.sheetnames:
        raise SystemExit(f"Missing required '{ANSWER_KEY_SHEET}' sheet in {path}")

    answer_key = {
        str(r["question_id"]).strip(): str(r["correct_option"]).strip().upper()
        for r in _read_sheet(wb[ANSWER_KEY_SHEET])
    }

    rows = []
    missing_answers = []
    for sheet_name, round_type in SHEET_TO_ROUND_TYPE.items():
        if sheet_name not in wb.sheetnames:
            continue
        for record in _read_sheet(wb[sheet_name]):
            qid = str(record["question_id"]).strip()
            if qid not in answer_key:
                missing_answers.append(qid)
                continue
            row = {key: record[key] for key in XLSX_QUESTION_COLUMNS}
            row["correct_option"] = answer_key[qid]
            row["round_type"] = round_type
            rows.append(row)

    if missing_answers:
        raise SystemExit(
            f"No entry in '{ANSWER_KEY_SHEET}' for question_id(s): {', '.join(missing_answers)}"
        )
    return rows


def load_bank(bank_path: str) -> int:
    path = Path(bank_path)
    if path.suffix.lower() == ".xlsx":
        rows = rows_from_xlsx(path)
    elif path.suffix.lower() == ".csv":
        rows = rows_from_csv(path)
    else:
        raise SystemExit(f"Unsupported file type: {path.suffix} (use .csv or .xlsx)")

    init_db()
    session = SessionLocal()
    session.query(Question).delete()
    for row in rows:
        session.add(
            Question(
                round_type=str(row["round_type"]).strip(),
                topic=str(row["topic"]).strip(),
                difficulty=str(row["difficulty"]).strip(),
                text=str(row["text"]).strip(),
                option_a=str(row["option_a"]).strip(),
                option_b=str(row["option_b"]).strip(),
                option_c=str(row["option_c"]).strip(),
                option_d=str(row["option_d"]).strip(),
                correct_option=str(row["correct_option"]).strip().upper(),
                points=int(row["points"]),
            )
        )
    session.commit()
    return len(rows)


def main(bank_path: str) -> None:
    count = load_bank(bank_path)
    print(f"Loaded {count} questions from {bank_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python load_questions.py <csv_or_xlsx_path>")
    main(sys.argv[1])
